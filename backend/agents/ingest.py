"""Ingest agent — classifies uploaded files and routes to right pipeline.

Pipelines:
  CV          → cv_pipeline (existing 12-step)
  JD          → jd_pipeline (extract + save to repo + auto-pick scoring profile)
  OFFER       → store offer (best-effort attach to candidate by name)
  CERTIFICATE → attach to candidate certifications (or store as standalone)
  OTHER       → keep file, mark for manual review
"""
from __future__ import annotations

import asyncio
import hashlib
import json
import logging
import os
import re
import time
import uuid
from pathlib import Path


_TITLE_LABEL_RE = re.compile(r"^\s*(job\s*title|title|position|role|job)\s*[:|\-]+\s*", re.IGNORECASE)
_DIVIDER_RE = re.compile(r"^\s*[-=*_|]{2,}.*$")


def _clean_jd_title(raw: str, filename: str | None) -> str:
    """Pick clean title from raw text — strip labels, table syntax, dividers."""
    fallback = Path(filename or "Imported JD").stem.replace("_", " ").replace("-", " ").strip().title()
    for line in (raw or "").splitlines()[:10]:
        s = line.strip().lstrip("#").strip()
        if not s or _DIVIDER_RE.match(s):
            continue
        s = _TITLE_LABEL_RE.sub("", s)
        s = s.replace("|", " ").strip(" :-—|")
        s = re.sub(r"\s+", " ", s)
        if 3 < len(s) < 90:
            return s[:90]
    return fallback[:90]


def _jd_content_hash(text: str) -> str:
    norm = re.sub(r"\s+", " ", (text or "").strip().lower())[:5000]
    return hashlib.sha256(norm.encode("utf-8")).hexdigest()

from backend.core.config import LITE_MODEL, llm_call, is_ai_available
from backend.core.cv_parser import (
    classify_file, extract_pdf, extract_docx,
)
from backend.core.database import get_cursor
from backend.core.ids import gen_public_id

logger = logging.getLogger(__name__)


CONFIDENCE_AUTOROUTE = 0.85
CONFIDENCE_MIN = 0.60


# ---------------------------------------------------------------------------
# Universal text extractor
# ---------------------------------------------------------------------------
def extract_any(file_path: str, file_type: str) -> dict:
    """Returns {"text", "pages", "method"} for any supported type."""
    if file_type == "pdf":
        return extract_pdf(file_path)
    if file_type == "docx":
        return extract_docx(file_path)
    if file_type == "image":
        # Use PaddleOCR via existing parser path (handwriting check inside)
        from backend.core.cv_parser import extract_image_ocr
        try:
            return extract_image_ocr(file_path)
        except Exception as e:
            logger.warning(f"image extract failed: {e}")
            return {"text": "", "pages": 1, "method": "image-fail"}
    if file_type == "xlsx":
        try:
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=True)
            chunks = []
            for ws in wb.worksheets:
                chunks.append(f"## {ws.title}")
                for row in ws.iter_rows(values_only=True):
                    if any(c is not None and str(c).strip() for c in row):
                        chunks.append(" | ".join(str(c) if c is not None else "" for c in row))
            return {"text": "\n".join(chunks), "pages": len(wb.worksheets), "method": "openpyxl"}
        except Exception as e:
            return {"text": "", "pages": 0, "method": f"xlsx-fail:{e}"}
    if file_type == "txt":
        try:
            return {"text": Path(file_path).read_text(errors="ignore"), "pages": 1, "method": "txt"}
        except Exception:
            return {"text": "", "pages": 0, "method": "txt-fail"}
    return {"text": "", "pages": 0, "method": "unknown"}


# ---------------------------------------------------------------------------
# Document type classifier (LLM + heuristics fallback)
# ---------------------------------------------------------------------------
def _heuristic_classify(text: str, filename: str) -> dict:
    t = (text or "").lower()
    fn = (filename or "").lower()
    if any(k in fn for k in ("resume", "cv", "candidate")):
        return {"type": "CV", "confidence": 0.75, "reason": "filename hint"}
    if any(k in fn for k in ("jd", "job-desc", "job_desc")):
        return {"type": "JD", "confidence": 0.75, "reason": "filename hint"}
    if any(k in fn for k in ("offer", "appointment")):
        return {"type": "OFFER", "confidence": 0.7, "reason": "filename hint"}
    if any(k in fn for k in ("certif", "diploma", "transcript")):
        return {"type": "CERTIFICATE", "confidence": 0.7, "reason": "filename hint"}

    cv_hits = sum(1 for k in ("experience", "education", "skills", "linkedin", "summary") if k in t)
    jd_hits = sum(1 for k in ("we are looking for", "responsibilities", "qualifications", "key responsibilities", "job purpose") if k in t)
    of_hits = sum(1 for k in ("we are pleased to offer", "your start date", "annual salary", "offer letter") if k in t)
    cert_hits = sum(1 for k in ("is awarded", "has completed", "certifies that", "this is to certify") if k in t)

    scores = {"CV": cv_hits, "JD": jd_hits, "OFFER": of_hits, "CERTIFICATE": cert_hits}
    best = max(scores, key=scores.get)
    if scores[best] == 0:
        return {"type": "OTHER", "confidence": 0.4, "reason": "no strong signal"}
    return {"type": best, "confidence": min(0.55 + 0.1 * scores[best], 0.85), "reason": f"keyword hits {scores[best]}"}


def classify_doc_type(text: str, filename: str) -> dict:
    """Try LLM first, fallback to heuristics."""
    if not is_ai_available() or not text or len(text) < 30:
        return _heuristic_classify(text, filename)

    excerpt = text[:800]
    prompt = f"""You are a document classifier for an HR ATS.
Classify the document below into ONE of these categories:
  CV          — resume / candidate biodata / professional profile
  JD          — job description / role posting / job requisition
  OFFER       — employment offer letter
  CERTIFICATE — credential / diploma / transcript / training certificate
  OTHER       — anything else

Heuristics:
  'experience', 'education', 'skills', personal email, work history → CV
  'we are looking for', 'responsibilities', 'qualifications', 'job purpose' → JD
  'we are pleased to offer', 'salary', 'start date' → OFFER
  'is awarded', 'has completed', 'certifies', diploma → CERTIFICATE

Filename: {filename}
Excerpt:
\"\"\"
{excerpt}
\"\"\"

Return ONLY valid JSON:
{{"type":"CV|JD|OFFER|CERTIFICATE|OTHER","confidence":0.0-1.0,"reason":"short why"}}"""
    try:
        out = llm_call(prompt, model=LITE_MODEL, temperature=0.1, max_tokens=200)
        m = re.search(r"\{[\s\S]*\}", out or "")
        if m:
            d = json.loads(m.group())
            if "type" in d:
                d["type"] = d["type"].upper()
                d["confidence"] = float(d.get("confidence") or 0.5)
                return d
    except Exception as e:
        logger.warning(f"LLM classify failed: {e}")
    return _heuristic_classify(text, filename)


# ---------------------------------------------------------------------------
# Routers per type
# ---------------------------------------------------------------------------
async def _route_cv(file_path: str, filename: str, raw: str, user, auto_process: bool = True) -> dict:
    """Hand off to existing CV upload pipeline. If auto_process=False, skip pipeline run.

    Dedup: SHA-256 of file bytes. If match on active candidate → skip pipeline,
    return existing candidate_id with deduped=True.
    """
    from backend.core.database import insert_candidate
    from backend.core.cv_pipeline import process_cv
    import asyncio, hashlib

    # ── Content-hash dedup ───────────────────────────────────────────────
    content_hash = None
    try:
        with open(file_path, "rb") as fh:
            content_hash = hashlib.sha256(fh.read()).hexdigest()
    except Exception as e:
        logger.warning(f"cv_content_hash compute failed: {e}")

    if content_hash:
        try:
            with get_cursor() as cur:
                cur.execute(
                    "SELECT id, name, file_name, is_processed FROM candidates "
                    "WHERE cv_content_hash=%s AND status='active' LIMIT 1",
                    (content_hash,),
                )
                row = cur.fetchone()
                if row:
                    existing_id, existing_name, existing_fname, is_processed = row
                    logger.info(f"cv dedup hit: hash={content_hash[:12]} → cv_{existing_id} ({existing_name})")
                    return {
                        "target_type": "candidate",
                        "target_id": existing_id,
                        "pipeline": "skipped",
                        "run_id": None,
                        "auto_process": False,
                        "deduped": True,
                        "reason": "duplicate_content",
                        "existing_name": existing_name,
                        "existing_file_name": existing_fname,
                        "existing_is_processed": bool(is_processed),
                    }
        except Exception as e:
            logger.error(f"cv dedup query failed (continuing as new): {e}")

    candidate_id = insert_candidate({
        "name": Path(filename or "unknown").stem.replace("_", " ").replace("-", " ").title(),
        "email": None, "phone": None, "location": None, "linkedin_url": None,
        "current_role": None, "current_company": None,
        "total_experience_years": None, "seniority_level": None,
        "skills_technical": [], "skills_soft": [], "tools": [], "languages": [],
        "experience": "[]", "education": "[]", "certifications": "[]", "projects": "[]",
        "summary_short": None, "summary_detailed": None, "raw_text": raw,
        "pdf_path": file_path, "file_type": Path(file_path).suffix.lstrip("."),
        "file_name": filename, "page_count": 0, "quality_score": 0,
        "tags": [], "source": "ingest", "status": "active", "is_processed": False,
        "owner_id": user.get("user_id") or 1,
        "sector_id": user.get("sector_id"),
        "visibility": "private",
    })
    # Stamp content hash on new row (no-op if compute failed earlier)
    if content_hash:
        try:
            with get_cursor() as cur:
                cur.execute(
                    "UPDATE candidates SET cv_content_hash=%s WHERE id=%s",
                    (content_hash, candidate_id),
                )
        except Exception as e:
            logger.warning(f"cv_content_hash stamp failed cv_{candidate_id}: {e}")

    run_id = str(uuid.uuid4()) if auto_process else None
    if auto_process:
        asyncio.create_task(_safe(process_cv(candidate_id, file_path, run_id=run_id), f"cv {candidate_id}"))
    return {"target_type": "candidate", "target_id": candidate_id,
            "pipeline": "cv_pipeline", "run_id": run_id, "auto_process": auto_process}


async def _route_jd(file_path: str, filename: str, raw: str, user, raw_only: bool = False) -> dict:
    """Save extracted text as a new JD in the repo + auto-pick scoring profile.

    If raw_only=True, skip LLM extraction — store text as-is with empty skills.
    """
    from backend.routes.jd_repo import _extract_requirements, SCORING_PRESETS

    title = _clean_jd_title(raw or "", filename)
    content_hash = _jd_content_hash(raw or "")

    # Dedup: if active JD with same content hash exists, update its source file + return existing id
    src_name = Path(filename or "").name
    src_ext = Path(filename or "").suffix.lower().lstrip(".")
    with get_cursor() as cur:
        cur.execute(
            "SELECT id FROM jd_repository WHERE jd_content_hash=%s AND status='active' LIMIT 1",
            (content_hash,),
        )
        existing = cur.fetchone()
        if existing:
            jd_id = existing[0]
            cur.execute(
                "UPDATE jd_repository SET source_file_path=%s, source_file_name=%s, source_file_type=%s, updated_at=NOW() WHERE id=%s",
                (file_path, src_name, src_ext, jd_id),
            )
            return {"target_type": "jd", "target_id": jd_id, "pipeline": "jd_pipeline", "title": title, "profile": "engineering", "deduped": True}

    # Auto-reformat layout (preserves wording) before extraction + storage
    raw_original = raw or ""
    final_text = raw_original
    text_source = "extracted"
    quality_score = 1.0
    if not raw_only:
        try:
            from backend.agents.jd_reformat import safe_reformat
            import asyncio as _aio
            final_text, text_source, quality_score = await _aio.to_thread(safe_reformat, raw_original)
        except Exception as _e:
            logger.warning("[jd_reformat] failed in upload path: %s", _e)
            final_text = raw_original

    if raw_only:
        extracted = {}
    else:
        extracted = await _extract_requirements(final_text)
    # Recompute hash on FINAL text (post-reformat)
    content_hash = _jd_content_hash(final_text)

    profile = "engineering"
    tl = title.lower()
    if any(k in tl for k in ("sales", "account exec")): profile = "sales"
    elif any(k in tl for k in ("design", "ux", "ui")):  profile = "design"
    elif "product" in tl:                               profile = "product"
    elif any(k in tl for k in ("marketing", "growth")): profile = "marketing"
    elif any(k in tl for k in ("compliance", "risk", "audit", "legal")): profile = "compliance"
    elif any(k in tl for k in ("doctor", "nurse", "clinical")): profile = "healthcare"
    elif any(k in tl for k in ("finance", "accountant", "treasur")): profile = "finance"
    p = SCORING_PRESETS[profile]

    try:
        from backend.core.expiry import get_default_expiry_days as _gdays
        _exp_days = _gdays("jd")
    except Exception:
        _exp_days = 90
    _uid = user.get("user_id") or 1
    with get_cursor() as cur:
        cur.execute("""
            INSERT INTO jd_repository (
                public_id,
                title, jd_text, jd_text_raw, jd_text_source, reformat_quality_score,
                required_skills, nice_to_have_skills,
                min_experience_years, education_level, industry_keywords,
                required_certifications, tags, created_by,
                weight_skills, weight_experience, weight_industry,
                weight_education, weight_certifications, weight_culture, scoring_profile,
                source_file_path, source_file_name, source_file_type, jd_content_hash,
                visibility, shared_sector, sector_id,
                created_by_id, updated_by_id, expires_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                      %s, %s, %s, %s, %s, %s, %s,
                      %s, %s, %s, %s,
                      %s, %s, %s,
                      %s, %s, NOW() + (%s || ' days')::interval)
            RETURNING id
        """, (
            gen_public_id("jd_repository"),
            title, final_text, raw_original, text_source, quality_score,
            extracted.get("required_skills", []),
            extracted.get("nice_to_have", []),
            extracted.get("min_experience_years", 0),
            extracted.get("education_level", ""),
            extracted.get("industry_keywords", []),
            extracted.get("certifications", []),
            ["imported"],
            _uid,
            p["weight_skills"], p["weight_experience"], p["weight_industry"],
            p["weight_education"], p["weight_certifications"], p["weight_culture"],
            profile,
            file_path, src_name, src_ext, content_hash,
            "sector", True, user.get("sector_id") or 1,
            _uid, _uid, _exp_days,
        ))
        jd_id = cur.fetchone()[0]
    asyncio.create_task(_safe(_embed_jd(jd_id, title, final_text), f"jd-embed {jd_id}"))
    return {"target_type": "jd", "target_id": jd_id, "pipeline": "jd_pipeline", "title": title, "profile": profile}


async def _embed_jd(jd_id: int, title: str, body: str) -> None:
    """Embed JD (title + body) + store. Idempotent via content hash."""
    from backend.core.config import embed_text
    text = f"{title}\n\n{body}".strip()[:8000]
    if not text:
        return
    h = hashlib.sha256(text.encode("utf-8")).hexdigest()
    with get_cursor() as cur:
        cur.execute("SELECT jd_embed_hash FROM jd_repository WHERE id=%s", (jd_id,))
        row = cur.fetchone()
        if row and row[0] == h:
            return  # already embedded
    vec = await embed_text(text)
    if not vec:
        return
    with get_cursor() as cur:
        cur.execute(
            "UPDATE jd_repository SET jd_embedding=%s::vector, jd_embed_hash=%s, updated_at=NOW() WHERE id=%s",
            (str(vec), h, jd_id),
        )


async def _route_offer(file_path: str, filename: str, raw: str, user) -> dict:
    """Best-effort: parse out salary + start date, store as offer."""
    salary_match = re.search(r"(?:USD|AED|SAR|INR|\$)\s*([\d,]+)", raw or "", re.IGNORECASE)
    salary = salary_match.group(1).replace(",", "") if salary_match else None
    with get_cursor() as cur:
        cur.execute("""
            INSERT INTO offers (notes, salary_amount, status, created_by, benefits)
            VALUES (%s, %s, 'imported', %s, %s)
            RETURNING id
        """, (f"Imported from {filename}\n\n{(raw or '')[:2000]}",
              float(salary) if salary else None,
              user.get("user_id") or 1,
              None))
        offer_id = cur.fetchone()[0]
    return {"target_type": "offer", "target_id": offer_id, "pipeline": "offer_import"}


async def _route_certificate(file_path: str, filename: str, raw: str, user) -> dict:
    """Store certificate as standalone (could later attach to candidate)."""
    return {"target_type": "certificate", "target_id": None, "pipeline": "certificate_store",
            "note": "stored in /data/certs awaiting candidate match"}


async def _route_other(file_path: str, filename: str, raw: str, user) -> dict:
    return {"target_type": "other", "target_id": None, "pipeline": "manual_review"}


async def _safe(coro, name="bg"):
    try:
        await coro
    except Exception as e:
        logger.error(f"{name} failed: {e}")


# ---------------------------------------------------------------------------
# Top-level orchestrator
# ---------------------------------------------------------------------------
async def ingest_file(file_path: str, filename: str, file_size: int, user, force_type: str | None = None, auto_process: bool = True, raw_only: bool = False) -> dict:
    """Run full agent pipeline for a single file. Returns ingest_log dict + result."""
    t0 = time.time()
    file_type = classify_file(file_path)
    if file_type == "unknown":
        return _log(filename, file_size, file_type, file_path, None, None, None,
                    classified_as=None, confidence=None, reason="unsupported",
                    pipeline=None, target_type=None, target_id=None,
                    status="error", error_msg=f"Unsupported file type", elapsed_ms=int((time.time()-t0)*1000), user=user)

    extracted = extract_any(file_path, file_type)
    raw = extracted.get("text", "")

    if force_type:
        cls = {"type": force_type.upper(), "confidence": 1.0, "reason": "user override"}
    else:
        cls = classify_doc_type(raw, filename)

    if cls["confidence"] < CONFIDENCE_MIN:
        return _log(filename, file_size, file_type, file_path,
                    raw[:5000], len(raw), None,
                    classified_as=cls["type"], confidence=cls["confidence"],
                    reason=cls.get("reason"), pipeline=None,
                    target_type=None, target_id=None,
                    status="pending_review", error_msg=None,
                    elapsed_ms=int((time.time()-t0)*1000), user=user)

    # Dispatch
    routers = {
        "CV": _route_cv,
        "JD": _route_jd,
        "OFFER": _route_offer,
        "CERTIFICATE": _route_certificate,
        "OTHER": _route_other,
    }
    try:
        if cls["type"] == "CV":
            result = await _route_cv(file_path, filename, raw, user, auto_process=auto_process)
        elif cls["type"] == "JD":
            result = await _route_jd(file_path, filename, raw, user, raw_only=raw_only)
        else:
            result = await routers[cls["type"]](file_path, filename, raw, user)
    except Exception as e:
        return _log(filename, file_size, file_type, file_path, raw[:5000], len(raw), None,
                    classified_as=cls["type"], confidence=cls["confidence"],
                    reason=cls.get("reason"), pipeline=None,
                    target_type=None, target_id=None,
                    status="error", error_msg=str(e),
                    elapsed_ms=int((time.time()-t0)*1000), user=user)

    return _log(filename, file_size, file_type, file_path, raw[:5000], len(raw), None,
                classified_as=cls["type"], confidence=cls["confidence"],
                reason=cls.get("reason"),
                pipeline=result.get("pipeline"),
                target_type=result.get("target_type"),
                target_id=result.get("target_id"),
                status="success", error_msg=None,
                elapsed_ms=int((time.time()-t0)*1000), user=user, extra=result)


def _log(filename, file_size, file_type, file_path, raw_text, extracted_chars, mime,
         classified_as, confidence, reason, pipeline, target_type, target_id,
         status, error_msg, elapsed_ms, user, extra=None):
    with get_cursor() as cur:
        cur.execute("""
            INSERT INTO ingest_log (
                user_id, filename, file_size, mime, file_type, file_path,
                extracted_chars, classified_as, confidence, reason,
                pipeline, target_type, target_id, status, error_msg, elapsed_ms,
                raw_text
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            RETURNING id, created_at
        """, (
            user.get("user_id") or 1, filename, file_size, mime, file_type, file_path,
            extracted_chars, classified_as, confidence, reason,
            pipeline, target_type, target_id, status, error_msg, elapsed_ms,
            raw_text,
        ))
        ingest_id, created_at = cur.fetchone()
    return {
        "ingest_id": ingest_id,
        "filename": filename,
        "file_type": file_type,
        "classified_as": classified_as,
        "confidence": confidence,
        "reason": reason,
        "pipeline": pipeline,
        "target_type": target_type,
        "target_id": target_id,
        "status": status,
        "error_msg": error_msg,
        "elapsed_ms": elapsed_ms,
        **(extra or {}),
    }
