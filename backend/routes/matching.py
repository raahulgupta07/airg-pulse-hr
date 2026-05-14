"""Matching routes — CV-JD scoring, ranking, auto-scan, compare."""
from __future__ import annotations

import json
import logging
import re
from typing import Optional

from fastapi import APIRouter, Depends, Request, HTTPException, Query

from backend.core.auth import get_current_user
from backend.core.permissions import require_role
from backend.core.database import (
    get_position, get_candidate, list_candidates,
    vector_search_candidates, vector_search_qa,
    add_candidate_to_position, get_position_candidates,
    get_cursor,
)
from backend.core.config import llm_call, embed_text, LITE_MODEL
from backend.routes.evaluation import generate_flags, score_culture

logger = logging.getLogger(__name__)
router = APIRouter()


# ---------------------------------------------------------------------------
# Fuzzy skill matching (Jaccard + substring)
# ---------------------------------------------------------------------------
_STOP = {
    "and", "or", "the", "of", "for", "to", "in", "on", "a", "an",
    "management", "system", "systems", "skills",
}
# Domain synonym buckets — collapse to canonical token so HSE ↔ EHS etc match.
_SYN: dict[str, str] = {
    "hse": "hse", "ehs": "hse", "eosh": "hse", "osh": "hse", "ohs": "hse",
    "ohsms": "hse", "ohsas": "hse",
    "environment": "environment", "environmental": "environment",
    "health": "health", "safety": "safety",
    "occupational": "occupational",
    "fire": "fire",
    "iso": "iso",
    "risk": "risk", "hazard": "risk",
    "audit": "audit", "auditing": "audit", "audits": "audit",
    "incident": "incident", "incidents": "incident",
    "rca": "rca", "root": "rca", "cause": "rca",
    "investigation": "investigation", "investigations": "investigation",
    "bcp": "bcp", "continuity": "bcp",
    "compliance": "compliance", "regulatory": "compliance",
    "training": "training", "trainings": "training",
    "myanmar": "myanmar",
    "chemical": "chemical", "chemicals": "chemical",
    "emergency": "emergency",
    "workplace": "workplace", "inspection": "inspection", "inspections": "inspection",
    "ppt": "powerpoint", "powerpoint": "powerpoint",
    "ms": "microsoft", "microsoft": "microsoft", "office": "office",
    "excel": "excel",
}


def _tokenize_skill(s: str) -> set[str]:
    if not s:
        return set()
    # Replace punctuation with space, lower
    cleaned = re.sub(r"[^a-zA-Z0-9]+", " ", s.lower())
    toks = [t for t in cleaned.split() if t and t not in _STOP and len(t) > 1]
    # Map synonyms
    return {_SYN.get(t, t) for t in toks}


def _skill_match(a: str, b: str, *, jaccard_thr: float = 0.4,
                 substr_thr: float = 0.6) -> bool:
    """True if two skill labels likely refer to the same concept."""
    if not a or not b:
        return False
    al, bl = a.lower().strip(), b.lower().strip()
    if al == bl:
        return True
    # Substring containment (e.g. "Fire Safety" in "Fire Safety Management")
    short, long = (al, bl) if len(al) <= len(bl) else (bl, al)
    if len(short) >= 4 and short in long and len(short) / len(long) >= substr_thr:
        return True
    ta, tb = _tokenize_skill(a), _tokenize_skill(b)
    if not ta or not tb:
        return False
    inter = ta & tb
    union = ta | tb
    if not union:
        return False
    jacc = len(inter) / len(union)
    if jacc >= jaccard_thr:
        return True
    # Asymmetric: if all of short-token-set is contained in long, count it
    smaller, bigger = (ta, tb) if len(ta) <= len(tb) else (tb, ta)
    if smaller and len(smaller & bigger) / len(smaller) >= 0.75:
        return True
    return False


# ---------------------------------------------------------------------------
# Scoring Functions
# ---------------------------------------------------------------------------
def score_skills(cv_skills: list, required: list, nice_to_have: list) -> dict:
    """Score skills overlap using fuzzy (Jaccard + substring) matching."""
    if not required:
        return {"score": 50, "matched": [], "missing": []}

    cv_list = [s for s in (cv_skills or []) if s]
    req_list = [s for s in required if s]
    nice_list = [s for s in (nice_to_have or []) if s]

    matched_req: list[str] = []
    missing_req: list[str] = []
    for req in req_list:
        hit = next((cv for cv in cv_list if _skill_match(req, cv)), None)
        if hit:
            matched_req.append(req)
        else:
            missing_req.append(req)

    matched_nice: list[str] = []
    for nice in nice_list:
        hit = next((cv for cv in cv_list if _skill_match(nice, cv)), None)
        if hit:
            matched_nice.append(nice)

    if req_list:
        req_pct = len(matched_req) / len(req_list) * 80
    else:
        req_pct = 80

    if nice_list:
        nice_pct = len(matched_nice) / len(nice_list) * 20
    else:
        nice_pct = 10

    score = min(100, round(req_pct + nice_pct))

    return {
        "score": score,
        "matched": matched_req + matched_nice,
        "missing": missing_req,
    }


def score_experience(cv_years: float, min_years: int) -> dict:
    """Score experience years fit."""
    if not min_years or min_years == 0:
        return {"score": 70, "detail": "No minimum specified"}

    cv_years = cv_years or 0

    if cv_years >= min_years * 1.5:
        score = 100  # Exceeds significantly
    elif cv_years >= min_years:
        score = 80 + (cv_years - min_years) / min_years * 20  # 80-100
    elif cv_years >= min_years * 0.7:
        score = 50 + (cv_years / min_years) * 30  # 50-80
    else:
        score = max(10, (cv_years / min_years) * 50)  # 10-50

    return {
        "score": min(100, round(score)),
        "detail": f"{cv_years}yr vs {min_years}yr required",
    }


def score_education(cv_education: list, required_level: str) -> dict:
    """Score education fit."""
    if not required_level:
        return {"score": 70, "detail": "No requirement"}

    level_rank = {"high_school": 1, "associate": 2, "bachelor": 3, "master": 4, "phd": 5, "doctorate": 5}
    req_rank = level_rank.get(required_level.lower(), 3)

    best_rank = 0
    for edu in (cv_education or []):
        degree = (edu.get("degree") or "").lower()
        for level, rank in level_rank.items():
            if level in degree:
                best_rank = max(best_rank, rank)

    if best_rank == 0:
        return {"score": 40, "detail": "Education not detected"}

    if best_rank >= req_rank:
        score = 80 + min(20, (best_rank - req_rank) * 10)
    else:
        score = max(20, 50 - (req_rank - best_rank) * 15)

    return {"score": min(100, round(score)), "detail": f"rank {best_rank} vs required {req_rank}"}


def score_certifications(cv_certs: list, required_certs: list) -> dict:
    """Score certification match."""
    if not required_certs:
        return {"score": 70, "detail": "No certifications required"}

    cv_cert_names = {(c.get("name") or "").lower() for c in (cv_certs or [])}
    req_names = {c.lower() for c in required_certs}

    matched = cv_cert_names & req_names
    score = len(matched) / len(req_names) * 100 if req_names else 70

    return {"score": min(100, round(score)), "matched": list(matched)}


def score_industry(cv_experience: list, industry_keywords: list) -> dict:
    """Score industry/domain relevance."""
    if not industry_keywords:
        return {"score": 70, "detail": "No industry keywords"}

    # Check company descriptions + roles against keywords
    exp_text = " ".join(
        f"{e.get('company', '')} {e.get('role', '')} {e.get('description', '')}"
        for e in (cv_experience or [])
    ).lower()

    keywords_lower = [k.lower() for k in industry_keywords]
    matched = [k for k in keywords_lower if k in exp_text]
    score = len(matched) / len(keywords_lower) * 100 if keywords_lower else 70

    return {"score": min(100, round(score)), "matched": matched}


def compute_composite(scores: dict, weights: dict) -> float:
    """Compute weighted composite score using all 6 dimensions.
    Auto-normalizes weights so composite stays in 0..100 even if input weights
    do not sum to exactly 100.
    """
    keys = ("skills", "experience", "industry", "education", "certifications", "culture", "competencies")
    total_w = sum(float(weights.get(k, 0) or 0) for k in keys)
    if total_w <= 0:
        return 0.0
    composite = 0.0
    for k in keys:
        w = float(weights.get(k, 0) or 0)
        s = float(scores.get(k, 0) or 0)
        composite += s * w / total_w
    return round(composite, 1)


def _persist_competency_score(position_id: int, candidate_id: int, comp_score: float, comp_gaps: list):
    """Persist competency score + gaps onto position_candidates (best-effort)."""
    try:
        from backend.core.database import get_cursor as _gc
        with _gc() as cur:
            cur.execute("""
                UPDATE position_candidates
                   SET match_score_competencies = %s,
                       competency_gaps = %s::jsonb,
                       updated_at = NOW()
                 WHERE position_id = %s AND candidate_id = %s
            """, (comp_score, json.dumps(comp_gaps or []), position_id, candidate_id))
    except Exception as e:
        logger.warning(f"persist competency score failed: {e}")


def _hydrate_position_from_jd(position: dict) -> dict:
    """If position lacks skills/keywords, pull from linked or title-matched JD repo row.
    Returns the (possibly augmented) position dict. Does NOT mutate DB."""
    needs = (
        not position.get("required_skills") or
        not position.get("nice_to_have_skills") or
        not position.get("industry_keywords") or
        not position.get("required_certifications")
    )
    if not needs and position.get("jd_embedding") is not None:
        return position
    try:
        from backend.core.database import get_cursor as _gc
        jd_id = position.get("weights_source_jd_id")
        with _gc() as cur:
            if jd_id:
                cur.execute("""
                    SELECT required_skills, nice_to_have_skills, industry_keywords,
                           required_certifications, min_experience_years, education_level,
                           jd_embedding
                    FROM jd_repository WHERE id = %s
                """, (jd_id,))
            else:
                cur.execute("""
                    SELECT required_skills, nice_to_have_skills, industry_keywords,
                           required_certifications, min_experience_years, education_level,
                           jd_embedding
                    FROM jd_repository
                    WHERE LOWER(title) = LOWER(%s) AND status = 'active'
                    ORDER BY updated_at DESC LIMIT 1
                """, (position.get("title") or "",))
            row = cur.fetchone()
            if not row:
                return position
            (req, nice, ind, certs, miny, edu, jd_emb) = row
            merged = dict(position)
            if not merged.get("required_skills"):
                merged["required_skills"] = req or []
            if not merged.get("nice_to_have_skills"):
                merged["nice_to_have_skills"] = nice or []
            if not merged.get("industry_keywords"):
                merged["industry_keywords"] = ind or []
            if not merged.get("required_certifications"):
                merged["required_certifications"] = certs or []
            if not merged.get("min_experience_years") and miny:
                merged["min_experience_years"] = miny
            if not merged.get("education_level") and edu:
                merged["education_level"] = edu
            if merged.get("jd_embedding") is None and jd_emb is not None:
                merged["jd_embedding"] = jd_emb
            return merged
    except Exception as e:
        logger.warning(f"hydrate from JD failed: {e}")
        return position


def _compute_semantic_score(position: dict, candidate_id: int) -> float | None:
    """Cosine similarity between JD embedding and CV full_cv embedding → 0-100.
    Returns None if either vector missing."""
    if not candidate_id:
        return None
    try:
        from backend.core.database import get_cursor as _gc
        with _gc() as cur:
            # Resolve JD embedding source: positions.jd_embedding first, then jd_repository.
            cur.execute("""
                SELECT (
                    SELECT (1 - (p.jd_embedding <=> ce.embedding))::float
                    FROM positions p
                    WHERE p.id = %s AND p.jd_embedding IS NOT NULL
                ) AS pos_sim,
                (
                    SELECT (1 - (jr.jd_embedding <=> ce.embedding))::float
                    FROM jd_repository jr
                    WHERE jr.id = COALESCE(
                        (SELECT weights_source_jd_id FROM positions WHERE id = %s),
                        (SELECT id FROM jd_repository
                         WHERE LOWER(title) = LOWER((SELECT title FROM positions WHERE id = %s))
                         ORDER BY updated_at DESC LIMIT 1)
                    ) AND jr.jd_embedding IS NOT NULL
                ) AS jd_sim
                FROM candidate_embeddings ce
                WHERE ce.candidate_id = %s AND ce.chunk_type = 'full_cv'
                ORDER BY ce.created_at DESC LIMIT 1
            """, (position.get("id"), position.get("id"), position.get("id"), candidate_id))
            row = cur.fetchone()
            if not row:
                return None
            pos_sim, jd_sim = row
            sim = pos_sim if pos_sim is not None else jd_sim
            if sim is None:
                return None
            # Cosine sim is in [-1, 1]; clamp + scale to 0-100.
            sim = max(0.0, min(1.0, float(sim)))
            return round(sim * 100, 1)
    except Exception as e:
        logger.warning(f"semantic score failed for candidate {candidate_id}: {e}")
        return None


def score_candidate_against_position(candidate: dict, position: dict) -> dict:
    """Full scoring of a candidate against a position."""
    # Hydrate JD-derived fields if position is missing them
    position = _hydrate_position_from_jd(position)

    skills_result = score_skills(
        candidate.get("skills_technical", []),
        position.get("required_skills", []),
        position.get("nice_to_have_skills", []),
    )
    exp_result = score_experience(
        candidate.get("total_experience_years"),
        position.get("min_experience_years", 0),
    )
    edu_result = score_education(
        candidate.get("education", []),
        position.get("education_level"),
    )
    cert_result = score_certifications(
        candidate.get("certifications", []),
        position.get("required_certifications", []),
    )
    ind_result = score_industry(
        candidate.get("experience", []),
        position.get("industry_keywords", []),
    )

    # Culture/values scoring
    culture_score = score_culture(candidate, position)

    # Competency scoring (gracefully no-op if no candidate_id / position_id)
    comp_score = 0.0
    comp_gaps: list = []
    try:
        if candidate.get("id") and position.get("id"):
            from backend.core.weights import score_competencies as _sc
            comp_result = _sc(candidate["id"], position["id"])
            comp_score = comp_result.get("score", 0)
            comp_gaps = comp_result.get("gaps", [])
            _persist_competency_score(position["id"], candidate["id"], comp_score, comp_gaps)
    except Exception as _e:
        logger.warning(f"competency scoring failed: {_e}")

    scores = {
        "skills": skills_result["score"],
        "experience": exp_result["score"],
        "education": edu_result["score"],
        "certifications": cert_result["score"],
        "industry": ind_result["score"],
        "culture": culture_score,
        "competencies": comp_score,
    }

    from backend.core.weights import effective_weights_for_scoring
    try:
        weights = effective_weights_for_scoring(position)
    except Exception:
        weights = {
            "skills": position.get("weight_skills", 40),
            "experience": position.get("weight_experience", 25),
            "education": position.get("weight_education", 10),
            "certifications": position.get("weight_certifications", 10),
            "industry": position.get("weight_industry", 15),
            "culture": position.get("weight_culture", 0),
            "competencies": position.get("weight_competencies", 0),
        }

    composite = compute_composite(scores, weights)

    # Semantic boost: cosine(JD embed, CV full_cv embed) blended at 20%
    semantic_score = _compute_semantic_score(position, candidate.get("id"))
    if semantic_score is not None:
        composite = round(composite * 0.8 + semantic_score * 0.2, 1)

    return {
        "composite": composite,
        "skills": skills_result["score"],
        "experience": exp_result["score"],
        "education": edu_result["score"],
        "certifications": cert_result["score"],
        "industry": ind_result["score"],
        "culture": culture_score,
        "competencies": comp_score,
        "semantic": semantic_score,
        "competency_gaps": comp_gaps,
        "weights_used": weights,
        "skills_matched": skills_result.get("matched", []),
        "skills_missing": skills_result.get("missing", []),
    }


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------
@router.post("/scan/{slug}")
async def scan_repo_for_position(slug: str, request: Request):
    """Scan entire central CV repo against a position's JD and recommend candidates."""
    user = get_current_user(request)

    position = get_position(slug)
    if not position:
        raise HTTPException(status_code=404, detail="Position not found")

    # Get all active candidates, excluding those already attached to this position
    from backend.core.database import get_cursor
    with get_cursor() as cur:
        cur.execute("""
            SELECT * FROM candidates
            WHERE status = 'active' AND is_processed = TRUE
              AND id NOT IN (SELECT candidate_id FROM position_candidates WHERE position_id = %s)
        """, (position["id"],))
        cols = [desc[0] for desc in cur.description]
        all_candidates = [dict(zip(cols, row)) for row in cur.fetchall()]

    if not all_candidates:
        return {"recommendations": [], "scanned": 0, "message": "All candidates already attached", "all_attached": True}

    # Score each candidate
    recommendations = []
    for candidate in all_candidates:
        # Parse JSONB fields
        for field in ("experience", "education", "certifications", "projects"):
            val = candidate.get(field)
            if isinstance(val, str):
                try:
                    candidate[field] = json.loads(val)
                except json.JSONDecodeError:
                    candidate[field] = []

        scores = score_candidate_against_position(candidate, position)

        # Skip below minimum threshold
        if scores["composite"] < position.get("min_match_score", 50):
            continue

        # Auto-generate flags
        try:
            generate_flags(position["id"], candidate["id"], candidate, position)
        except Exception as e:
            logger.warning(f"Flag generation failed for candidate {candidate['id']}: {e}")

        recommendations.append({
            "candidate_id": candidate["id"],
            "name": candidate["name"],
            "current_role": candidate.get("current_role"),
            "current_company": candidate.get("current_company"),
            "total_experience_years": candidate.get("total_experience_years"),
            "seniority_level": candidate.get("seniority_level"),
            "skills_technical": candidate.get("skills_technical", []),
            "scores": scores,
        })

    # Sort by composite score
    recommendations.sort(key=lambda x: x["scores"]["composite"], reverse=True)

    # Generate explanations for top 10
    for rec in recommendations[:10]:
        try:
            explanation = await _generate_match_explanation(rec, position)
            rec["explanation"] = explanation
        except Exception:
            rec["explanation"] = None

    return {
        "recommendations": recommendations[:50],
        "scanned": len(all_candidates),
        "above_threshold": len(recommendations),
        "threshold": position.get("min_match_score", 50),
    }


@router.post("/scan/{slug}/add", dependencies=[Depends(require_role("recruiter"))])
async def add_scan_results(slug: str, request: Request):
    """Add recommended candidates to a position."""
    user = get_current_user(request)
    body = await request.json()

    position = get_position(slug)
    if not position:
        raise HTTPException(status_code=404, detail="Position not found")

    candidate_ids = body.get("candidate_ids", [])
    # source: 'pool_pick' (Talent Pool attach), 'position_upload' (direct upload),
    # 'ai_scan'/'ai_promoted' (AI rec promote). Default 'manual' for legacy.
    source = (body.get("source") or "manual").lower()
    if source not in ("manual", "ai_scan", "ai_promoted", "pool_pick", "position_upload"):
        source = "manual"

    # Dedup pre-check: find which ids are already attached (any stage, including dismissed)
    already_attached: dict[int, dict] = {}
    if candidate_ids:
        try:
            with get_cursor() as cur:
                cur.execute(
                    "SELECT pc.candidate_id, c.name, pc.stage, pc.match_source, pc.dismissed "
                    "FROM position_candidates pc JOIN candidates c ON c.id=pc.candidate_id "
                    "WHERE pc.position_id=%s AND pc.candidate_id = ANY(%s)",
                    (position["id"], list(candidate_ids)),
                )
                for cid_row, name, stage, msrc, dismissed in cur.fetchall():
                    already_attached[int(cid_row)] = {
                        "candidate_id": int(cid_row),
                        "name": name,
                        "stage": stage,
                        "match_source": msrc,
                        "dismissed": bool(dismissed),
                    }
        except Exception as e:
            logger.warning(f"[scan/add] dedup pre-check failed: {e}")

    added = 0
    skipped: list[dict] = []
    not_found: list[int] = []

    for cid in candidate_ids:
        if int(cid) in already_attached:
            info = already_attached[int(cid)]
            skipped.append({
                **info,
                "reason": (
                    "already_rejected" if info["dismissed"] or info["stage"] == "rejected"
                    else "already_attached"
                ),
            })
            continue
        candidate = get_candidate(cid)
        if not candidate:
            not_found.append(int(cid))
            continue

        # Parse JSONB
        for field in ("experience", "education", "certifications"):
            val = candidate.get(field)
            if isinstance(val, str):
                try:
                    candidate[field] = json.loads(val)
                except json.JSONDecodeError:
                    candidate[field] = []

        scores = score_candidate_against_position(candidate, position)

        try:
            generate_flags(position["id"], cid, candidate, position)
        except Exception as e:
            logger.warning(f"Flag generation failed: {e}")

        add_candidate_to_position(
            position["id"], cid,
            scores={
                "composite": scores["composite"],
                "skills": scores["skills"],
                "experience": scores["experience"],
                "education": scores["education"],
                "certifications": scores["certifications"],
                "industry": scores["industry"],
                "culture": scores.get("culture"),
                "semantic": scores.get("semantic"),
                "explanation": None,
                "skills_matched": scores.get("skills_matched", []),
                "skills_missing": scores.get("skills_missing", []),
            },
            added_by=source,
        )
        added += 1

    return {
        "added": added,
        "skipped": skipped,
        "skipped_count": len(skipped),
        "not_found": not_found,
        "requested": len(candidate_ids),
        "message": (
            f"Added {added} of {len(candidate_ids)} (skipped {len(skipped)} dupes"
            + (f", {len(not_found)} not found" if not_found else "")
            + ")"
        ),
    }


@router.post("/score")
async def score_single(request: Request):
    """Score a single candidate against a position."""
    user = get_current_user(request)
    body = await request.json()

    candidate_id = body.get("candidate_id")
    position_slug = body.get("position_slug")

    candidate = get_candidate(candidate_id)
    position = get_position(position_slug)
    if not candidate or not position:
        raise HTTPException(status_code=404, detail="Candidate or position not found")

    for field in ("experience", "education", "certifications"):
        val = candidate.get(field)
        if isinstance(val, str):
            try:
                candidate[field] = json.loads(val)
            except json.JSONDecodeError:
                candidate[field] = []

    scores = score_candidate_against_position(candidate, position)
    return {"candidate_id": candidate_id, "position_slug": position_slug, "scores": scores}


@router.post("/compare")
async def compare_candidates(request: Request):
    """Compare candidates side-by-side. Position scoring optional — when no position_slug,
    returns raw attribute comparison without weighted scores.
    """
    user = get_current_user(request)
    body = await request.json()

    candidate_ids = body.get("candidate_ids", [])
    position_slug = body.get("position_slug")

    if len(candidate_ids) < 2 or len(candidate_ids) > 5:
        raise HTTPException(status_code=400, detail="Compare 2-5 candidates")

    position = get_position(position_slug) if position_slug else None
    if position_slug and not position:
        raise HTTPException(status_code=404, detail="Position not found")

    results = []
    for cid in candidate_ids:
        candidate = get_candidate(cid)
        if not candidate:
            continue

        for field in ("experience", "education", "certifications"):
            val = candidate.get(field)
            if isinstance(val, str):
                try:
                    candidate[field] = json.loads(val)
                except json.JSONDecodeError:
                    candidate[field] = []

        scores = score_candidate_against_position(candidate, position) if position else None
        results.append({
            "candidate_id": cid,
            "name": candidate.get("name"),
            "email": candidate.get("email"),
            "phone": candidate.get("phone"),
            "location": candidate.get("location"),
            "current_role": candidate.get("current_role"),
            "current_company": candidate.get("current_company"),
            "seniority_level": candidate.get("seniority_level"),
            "total_experience_years": candidate.get("total_experience_years"),
            "skills_technical": candidate.get("skills_technical", []),
            "skills_soft": candidate.get("skills_soft", []),
            "tools": candidate.get("tools", []),
            "certifications": candidate.get("certifications", []),
            "education": candidate.get("education", []),
            "experience": candidate.get("experience", []),
            "languages": candidate.get("languages", []),
            "quality_score": candidate.get("quality_score"),
            "ai_summary": candidate.get("ai_summary"),
            "scores": scores,
        })

    if position:
        results.sort(key=lambda x: (x.get("scores") or {}).get("composite", 0), reverse=True)
    else:
        results.sort(key=lambda x: x.get("quality_score") or 0, reverse=True)

    comparison_summary = None
    if len(results) >= 2:
        try:
            if position:
                comparison_summary = await _generate_comparison(results, position)
            else:
                comparison_summary = await _generate_open_comparison(results)
        except Exception as e:
            logger.warning(f"compare summary failed: {e}")

    return {
        "candidates": results,
        "comparison_summary": comparison_summary,
        "position": position_slug,
        "scored": bool(position),
    }


@router.get("/compare/export.xlsx")
async def compare_export_xlsx(request: Request, ids: str = ""):
    """Export comparison as Excel (.xlsx). ids=comma-separated candidate ids."""
    get_current_user(request)
    try:
        cid_list = [int(x.strip()) for x in ids.split(",") if x.strip()]
    except ValueError:
        raise HTTPException(400, "Invalid ids")
    if len(cid_list) < 1 or len(cid_list) > 10:
        raise HTTPException(400, "1-10 candidates per export")
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")
    import io as _io
    cands = []
    for cid in cid_list:
        c = get_candidate(cid)
        if not c: continue
        for field in ("experience", "education", "certifications"):
            v = c.get(field)
            if isinstance(v, str):
                try: c[field] = json.loads(v)
                except Exception: c[field] = []
        cands.append(c)
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"
    header_fill = PatternFill("solid", fgColor="383832")
    header_font = Font(bold=True, color="FEFFD6", size=11)
    dim_font = Font(bold=True, size=10, color="383832")
    border = Border(*[Side(style="thin", color="888888")] * 4)
    align_top = Alignment(vertical="top", wrap_text=True)
    # Row 1: headers
    headers = ["Dimension"] + [c.get("name", f"cand_{c.get('id')}") for c in cands]
    for col_i, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_i, value=h)
        cell.fill = header_fill; cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    dims = [
        ("Email", lambda c: c.get("email") or ""),
        ("Phone", lambda c: c.get("phone") or ""),
        ("Location", lambda c: c.get("location") or ""),
        ("Current Role", lambda c: c.get("current_role") or ""),
        ("Current Company", lambda c: c.get("current_company") or ""),
        ("Seniority", lambda c: (c.get("seniority_level") or "").upper()),
        ("Experience (yrs)", lambda c: c.get("total_experience_years") or 0),
        ("Quality Score", lambda c: c.get("quality_score") or 0),
        ("Technical Skills", lambda c: ", ".join(c.get("skills_technical") or [])),
        ("Soft Skills", lambda c: ", ".join(c.get("skills_soft") or [])),
        ("Tools", lambda c: ", ".join(c.get("tools") or [])),
        ("Languages", lambda c: ", ".join(str(x) for x in (c.get("languages") or []))),
        ("Certifications", lambda c: ", ".join(
            (x.get("name") if isinstance(x, dict) else str(x)) for x in (c.get("certifications") or []))),
        ("Education", lambda c: " | ".join(
            f"{e.get('degree','')} {e.get('field','')} @ {e.get('institution','')}"
            for e in (c.get("education") or []) if isinstance(e, dict))),
        ("Experience History", lambda c: " | ".join(
            f"{e.get('role','')} @ {e.get('company','')} ({e.get('start_date','')}-{e.get('end_date','')})"
            for e in (c.get("experience") or [])[:5] if isinstance(e, dict))),
        ("AI Summary", lambda c: c.get("ai_summary") or ""),
    ]
    for row_i, (label, fn) in enumerate(dims, start=2):
        ws.cell(row=row_i, column=1, value=label).font = dim_font
        for col_i, c in enumerate(cands, start=2):
            try: val = fn(c)
            except Exception: val = ""
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.alignment = align_top
            cell.border = border
    # Column widths
    ws.column_dimensions["A"].width = 22
    for col_i in range(2, 2 + len(cands)):
        ws.column_dimensions[chr(64 + col_i)].width = 45
    ws.row_dimensions[1].height = 22
    # Generate AI summary on a second sheet
    try:
        from openpyxl.utils import get_column_letter
        ws2 = wb.create_sheet("AI Summary")
        ws2["A1"] = "AI Executive Comparison"
        ws2["A1"].font = Font(bold=True, size=14, color="007518")
        cand_brief = [{
            "name": c.get("name"), "current_role": c.get("current_role"),
            "current_company": c.get("current_company"),
            "total_experience_years": c.get("total_experience_years"),
            "seniority_level": c.get("seniority_level"),
            "skills_technical": c.get("skills_technical") or [],
            "certifications": c.get("certifications") or [],
            "quality_score": c.get("quality_score"),
        } for c in cands]
        summary = await _generate_open_comparison(cand_brief) if len(cand_brief) >= 2 else None
        ws2["A3"] = summary or "(AI summary unavailable)"
        ws2["A3"].alignment = Alignment(wrap_text=True, vertical="top")
        ws2.column_dimensions["A"].width = 120
        ws2.row_dimensions[3].height = 240
    except Exception as e:
        logger.warning(f"xlsx ai summary failed: {e}")
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from fastapi.responses import StreamingResponse
    fname = "pulse-comparison.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.post("/{slug}/rescore", dependencies=[Depends(require_role("recruiter"))])
async def rescore_all(slug: str, request: Request):
    """Rescore all candidates for a position (after weight changes)."""
    user = get_current_user(request)

    position = get_position(slug)
    if not position:
        raise HTTPException(status_code=404, detail="Position not found")

    candidates = get_position_candidates(position["id"], limit=500)
    rescored = 0

    for pc in candidates:
        candidate = get_candidate(pc["candidate_id"])
        if not candidate:
            continue

        for field in ("experience", "education", "certifications"):
            val = candidate.get(field)
            if isinstance(val, str):
                try:
                    candidate[field] = json.loads(val)
                except json.JSONDecodeError:
                    candidate[field] = []

        scores = score_candidate_against_position(candidate, position)

        try:
            generate_flags(position["id"], pc["candidate_id"], candidate, position)
        except Exception as e:
            logger.warning(f"Flag generation failed for candidate {pc['candidate_id']}: {e}")

        add_candidate_to_position(
            position["id"], pc["candidate_id"],
            scores={
                "composite": scores["composite"],
                "skills": scores["skills"],
                "experience": scores["experience"],
                "education": scores["education"],
                "certifications": scores["certifications"],
                "industry": scores["industry"],
                "culture": scores.get("culture"),
                "semantic": scores.get("semantic"),
                "explanation": None,
                "skills_matched": scores.get("skills_matched", []),
                "skills_missing": scores.get("skills_missing", []),
            },
            added_by=pc.get("added_by", "manual"),
        )
        rescored += 1

    return {"rescored": rescored, "message": f"Rescored {rescored} candidates"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
async def _generate_match_explanation(rec: dict, position: dict) -> str | None:
    """LLM-generate why this candidate matches."""
    prompt = f"""In 2-3 sentences, explain why this candidate is a good match for this position.

Position: {position.get('title')}
Required Skills: {', '.join(position.get('required_skills', []))}
Min Experience: {position.get('min_experience_years', 0)} years

Candidate: {rec.get('name')}
Role: {rec.get('current_role')} at {rec.get('current_company')}
Experience: {rec.get('total_experience_years')} years
Skills: {', '.join(rec.get('skills_technical', [])[:10])}
Match Score: {rec['scores']['composite']}%

Be specific about strengths and any gaps."""

    return llm_call(prompt, model=LITE_MODEL, temperature=0.3, max_tokens=200)


async def _generate_comparison(candidates: list, position: dict) -> str | None:
    """LLM-generate comparison summary."""
    cand_lines = []
    for c in candidates:
        cand_lines.append(
            f"- {c['name']}: {c['scores']['composite']}% overall, "
            f"Skills:{c['scores']['skills']}% Exp:{c['scores']['experience']}%"
        )

    prompt = f"""Compare these candidates for the {position.get('title')} position.
Write a 3-4 sentence summary identifying the strongest candidate and key differentiators.

{chr(10).join(cand_lines)}"""

    return llm_call(prompt, model=LITE_MODEL, temperature=0.3, max_tokens=300)


async def _generate_open_comparison(candidates: list) -> str | None:
    """LLM-generate comparison without a specific position — focus on profile strengths."""
    lines = []
    for c in candidates:
        skills = ", ".join((c.get("skills_technical") or [])[:8])
        certs_raw = c.get("certifications") or []
        if isinstance(certs_raw, list) and certs_raw and isinstance(certs_raw[0], dict):
            certs = ", ".join(x.get("name", "") for x in certs_raw[:5])
        else:
            certs = ", ".join(str(x) for x in (certs_raw or [])[:5])
        lines.append(
            f"- {c.get('name','?')} | {c.get('current_role','?')} @ {c.get('current_company','?')} | "
            f"{c.get('total_experience_years',0)}y | {c.get('seniority_level','?')} | "
            f"Q={c.get('quality_score',0)} | Skills: {skills} | Certs: {certs}"
        )
    prompt = f"""You are an expert HR analyst. Compare these candidates side-by-side and produce a concise executive brief (4-6 sentences).
Cover: (1) shared strengths, (2) distinct strengths per candidate, (3) gaps to probe in interview, (4) which role-type each suits best, (5) one-line hiring recommendation.

CANDIDATES:
{chr(10).join(lines)}

Write in plain text, no markdown headers, no bullets — flowing prose. End with a clear recommendation line starting with "RECOMMENDATION:"."""
    return llm_call(prompt, model=LITE_MODEL, temperature=0.3, max_tokens=500)


# ---------------------------------------------------------------------------
# AI Recommendations — extends scan with assignments + status per candidate
# ---------------------------------------------------------------------------
from pydantic import BaseModel as _BM


@router.get("/recommendations/{slug}")
async def recommendations(
    slug: str,
    request: Request,
    available_only: bool = False,
    min_score: int = 60,
):
    """Return scored candidates for a position, with each candidate's other
    assignments inline + status badge (available / in_use / in_this / dismissed).
    """
    user = get_current_user(request)
    from backend.core.database import get_cursor as _gc

    with _gc() as cur:
        cur.execute("SELECT id FROM positions WHERE slug = %s", (slug,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Position not found")
        position_id = row[0]

    # Run the existing scan (returns scored recommendations)
    scan_result = await scan_repo_for_position(slug, request)
    recs = scan_result.get("recommendations") or []

    # Filter by min_score
    recs = [r for r in recs if (r.get("score") or r.get("scores", {}).get("composite", 0)) >= min_score]

    # Pull current assignments + dismissal flags for these candidates
    cand_ids = [r.get("candidate_id") or r.get("id") for r in recs if r.get("candidate_id") or r.get("id")]
    if not cand_ids:
        return {"recommendations": [], "position_id": position_id}

    with _gc() as cur:
        cur.execute("""
            SELECT pc.candidate_id, pc.position_id, p.slug, p.title, pc.stage,
                   pc.match_score_composite,
                   pc.created_at AS added_at,
                   pc.added_by AS recruiter_name,
                   COALESCE(pc.dismissed, FALSE) AS dismissed
            FROM position_candidates pc
            JOIN positions p ON p.id = pc.position_id
            WHERE pc.candidate_id = ANY(%s)
        """, (cand_ids,))
        cols = [d[0] for d in cur.description]
        all_assignments = [dict(zip(cols, r)) for r in cur.fetchall()]

    by_cand: dict[int, list] = {}
    dismissed_for_this: set[int] = set()
    in_this: set[int] = set()
    for a in all_assignments:
        cid = a["candidate_id"]
        if a["position_id"] == position_id:
            if a["dismissed"]:
                dismissed_for_this.add(cid)
            else:
                in_this.add(cid)
            continue
        if a["dismissed"]:
            continue
        by_cand.setdefault(cid, []).append({
            "position_id": a["position_id"],
            "slug": a["slug"],
            "title": a["title"],
            "stage": a["stage"],
            "match_score": a["match_score_composite"],
            "recruiter_name": a["recruiter_name"],
            "added_at": a["added_at"].isoformat() if a["added_at"] else None,
        })

    # Build response
    out = []
    for r in recs:
        cid = r.get("candidate_id") or r.get("id")
        assignments = by_cand.get(cid, [])
        if cid in in_this:
            status = "in_this"
        elif cid in dismissed_for_this:
            status = "dismissed"
        elif assignments:
            status = "in_use"
        else:
            status = "available"
        if available_only and status != "available":
            continue
        out.append({
            **r,
            "candidate_id": cid,
            "assignments": assignments,
            "status": status,
        })

    # Dedup by (lowercase name, lowercase email) — same person uploaded twice
    # → keep highest-scoring record, merge assignments + duplicate_ids
    deduped: dict = {}
    for r in out:
        key = ((r.get("name") or "").strip().lower(), (r.get("email") or "").strip().lower())
        if not key[0] and not key[1]:
            # No identity — keep as-is
            deduped[("__noid__", r.get("candidate_id"))] = r
            continue
        existing = deduped.get(key)
        if existing is None:
            r = {**r, "duplicate_candidate_ids": []}
            deduped[key] = r
            continue
        # Merge: keep higher composite, append assignments + dup id
        cur_score = (existing.get("scores") or {}).get("composite") or existing.get("score") or 0
        new_score = (r.get("scores") or {}).get("composite") or r.get("score") or 0
        winner, loser = (existing, r) if cur_score >= new_score else (r, existing)
        merged_assignments = (winner.get("assignments") or []) + (loser.get("assignments") or [])
        seen_a = set()
        deduped_assignments = []
        for a in merged_assignments:
            key_a = (a.get("position_id"), a.get("stage"))
            if key_a in seen_a:
                continue
            seen_a.add(key_a)
            deduped_assignments.append(a)
        winner = {**winner, "assignments": deduped_assignments}
        winner.setdefault("duplicate_candidate_ids", [])
        winner["duplicate_candidate_ids"] = list(set(
            (existing.get("duplicate_candidate_ids") or []) +
            (r.get("duplicate_candidate_ids") or []) +
            [loser.get("candidate_id")]
        ))
        deduped[key] = winner
    out = list(deduped.values())

    return {
        "recommendations": out,
        "position_id": position_id,
        "total": len(out),
        "min_score": min_score,
    }


class DismissRequest(_BM):
    candidate_ids: list[int]
    reason: Optional[str] = None


@router.post("/scan/{slug}/dismiss", dependencies=[Depends(require_role("recruiter"))])
async def dismiss_recommendations(slug: str, body: DismissRequest, request: Request):
    user = get_current_user(request)
    from backend.core.database import get_cursor as _gc
    import json as _json
    with _gc() as cur:
        cur.execute("SELECT id FROM positions WHERE slug = %s", (slug,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Position not found")
        position_id = row[0]

        # Insert if not exists (so we can mark dismissed for never-attached candidates)
        for cid in body.candidate_ids:
            cur.execute("""
                INSERT INTO position_candidates (position_id, candidate_id, stage, added_by, ai_recommended, dismissed, dismissed_by, dismissed_at, dismissal_reason)
                VALUES (%s, %s, 'uploaded', %s, TRUE, TRUE, %s, NOW(), %s)
                ON CONFLICT (position_id, candidate_id) DO UPDATE SET
                    dismissed = TRUE,
                    dismissed_by = EXCLUDED.dismissed_by,
                    dismissed_at = EXCLUDED.dismissed_at,
                    dismissal_reason = EXCLUDED.dismissal_reason
            """, (position_id, cid, str(user.get("user_id")), user.get("user_id"), body.reason))
        cur.execute(
            "INSERT INTO audit_log (user_id, action, resource_type, resource_id, details) VALUES (%s,%s,%s,%s,%s)",
            (user.get("user_id"), "AI_REC_DISMISS", "position_candidates", position_id,
             _json.dumps({"slug": slug, "candidate_ids": body.candidate_ids, "reason": body.reason})),
        )
    return {"dismissed": len(body.candidate_ids), "position_id": position_id}


@router.post("/scan/{slug}/undismiss", dependencies=[Depends(require_role("recruiter"))])
async def undismiss_recommendations(slug: str, body: DismissRequest, request: Request):
    user = get_current_user(request)
    from backend.core.database import get_cursor as _gc
    with _gc() as cur:
        cur.execute("SELECT id FROM positions WHERE slug = %s", (slug,))
        row = cur.fetchone()
        if not row:
            raise HTTPException(404, "Position not found")
        position_id = row[0]
        cur.execute("""
            UPDATE position_candidates
               SET dismissed = FALSE, dismissed_by = NULL,
                   dismissed_at = NULL, dismissal_reason = NULL
             WHERE position_id = %s AND candidate_id = ANY(%s)
        """, (position_id, body.candidate_ids))
    return {"undismissed": len(body.candidate_ids), "position_id": position_id}
