"""Export vision bench results to Excel."""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
RUNS = ROOT / "bench" / "runs"
GT = json.load(open(ROOT / "bench" / "ground_truth.json"))
OUT = ROOT / "bench" / "vision_results.xlsx"

green = PatternFill("solid", fgColor="C6EFCE")
red = PatternFill("solid", fgColor="FFC7CE")
hdr_fill = PatternFill("solid", fgColor="383832")
hdr_font = Font(bold=True, color="FEFFD6", size=11)
title_font = Font(bold=True, size=14)
border = Border(*[Side(style="thin", color="888888")] * 4)

# Load runs
runs = []
for f in sorted(RUNS.glob("*.json")):
    r = json.load(open(f))
    if r.get("ok"): runs.append(r)

wb = Workbook()

# === Sheet 1: SUMMARY ===
ws = wb.active
ws.title = "Summary"
ws["A1"] = "Vision Model Bench — Myanmar Driver Form (1 doc)"
ws["A1"].font = title_font
ws.merge_cells("A1:G1")

cols = ["Model", "Provider", "Field Acc", "Acc %", "Latency (ms)", "Cost / doc (USD)", "Cost / 1000 docs (USD)"]
for i, c in enumerate(cols, 1):
    cell = ws.cell(row=3, column=i, value=c)
    cell.fill = hdr_fill; cell.font = hdr_font; cell.alignment = Alignment(horizontal="center"); cell.border = border

# Sort by cost ascending within accuracy descending
runs_sorted = sorted(runs, key=lambda r: (-r["score"]["field_acc"], r["cost_usd"]))
for i, r in enumerate(runs_sorted, 4):
    provider = r["model"].split("/")[0]
    acc = r["score"]["field_acc"]
    row = [r["model"], provider, f"{r['score']['n_match']}/{r['score']['n_total']}",
           f"{100*acc:.0f}%", r["latency_ms"], r["cost_usd"], round(r["cost_usd"]*1000, 4)]
    for j, v in enumerate(row, 1):
        c = ws.cell(row=i, column=j, value=v); c.border = border
        if j == 4:
            c.fill = green if acc == 1.0 else (red if acc < 0.9 else PatternFill("solid", fgColor="FFEB9C"))

for col in range(1, 8):
    ws.column_dimensions[get_column_letter(col)].width = [38, 12, 11, 9, 14, 18, 22][col-1]
ws.row_dimensions[1].height = 22

# === Sheet 2: FIELD-BY-FIELD ===
ws2 = wb.create_sheet("Per-field")
ws2["A1"] = "Per-field accuracy ✅ = match (fuzzy 0.80) ❌ = miss"
ws2["A1"].font = title_font
ws2.merge_cells("A1:Z1")

# header row: field + each model name
fields = list(GT.keys())
ws2.cell(row=3, column=1, value="FIELD").fill = hdr_fill
ws2.cell(row=3, column=1).font = hdr_font
ws2.cell(row=3, column=2, value="GROUND TRUTH").fill = hdr_fill
ws2.cell(row=3, column=2).font = hdr_font
for i, r in enumerate(runs_sorted, 3):
    c = ws2.cell(row=3, column=i, value=r["model"].split("/")[-1])
    c.fill = hdr_fill; c.font = hdr_font; c.alignment = Alignment(horizontal="center", wrap_text=True)

for fi, f in enumerate(fields, 4):
    ws2.cell(row=fi, column=1, value=f).font = Font(bold=True)
    ws2.cell(row=fi, column=2, value=str(GT[f]))
    for ri, r in enumerate(runs_sorted, 3):
        m = r["score"]["matches"][f]
        got = m.get("got") or "—"
        cell = ws2.cell(row=fi, column=ri, value=str(got))
        cell.fill = green if m["match"] else red
        cell.alignment = Alignment(wrap_text=True, vertical="top")

ws2.column_dimensions["A"].width = 18
ws2.column_dimensions["B"].width = 30
for i in range(3, 3 + len(runs_sorted)):
    ws2.column_dimensions[get_column_letter(i)].width = 28
ws2.row_dimensions[3].height = 30

# === Sheet 3: CRITICAL FIELDS ===
ws3 = wb.create_sheet("Critical fields")
ws3["A1"] = "Critical-field exact match (DOB + phone + NRC last digit)"
ws3["A1"].font = title_font
ws3.merge_cells("A1:F1")

crit = ["dob", "phone", "nrc"]
hdr = ["Model", "DOB exact", "Phone exact", "NRC exact", "Critical hits", "Cost/doc"]
for i, h in enumerate(hdr, 1):
    c = ws3.cell(row=3, column=i, value=h); c.fill = hdr_fill; c.font = hdr_font; c.border = border

def exact(got, gt):
    if got is None: return False
    return str(got).replace(" ", "").lower() == str(gt).replace(" ", "").lower()

for i, r in enumerate(runs_sorted, 4):
    hits = 0
    row_data = [r["model"]]
    for k in crit:
        got = r["score"]["matches"][k].get("got")
        ok = exact(got, GT[k])
        if ok: hits += 1
        row_data.append("✅" if ok else f"❌ {got}")
    row_data.append(f"{hits}/3")
    row_data.append(f"${r['cost_usd']:.5f}")
    for j, v in enumerate(row_data, 1):
        c = ws3.cell(row=i, column=j, value=v); c.border = border
        if 2 <= j <= 4:
            c.fill = green if v == "✅" else red
        if j == 5:
            c.fill = green if hits == 3 else (PatternFill("solid", fgColor="FFEB9C") if hits >= 2 else red)

for col, w in enumerate([38, 14, 35, 35, 14, 12], 1):
    ws3.column_dimensions[get_column_letter(col)].width = w

# === Sheet 4: COST PROJECTIONS ===
ws4 = wb.create_sheet("Cost projection")
ws4["A1"] = "Cost projection by volume"
ws4["A1"].font = title_font
ws4.merge_cells("A1:F1")

vols = [100, 1_000, 10_000, 100_000, 1_000_000]
hdr2 = ["Model", "Field acc"] + [f"${v:,} docs" for v in vols]
hdr2 = ["Model", "Field acc"] + [f"{v:,} docs" for v in vols]
for i, h in enumerate(hdr2, 1):
    c = ws4.cell(row=3, column=i, value=h); c.fill = hdr_fill; c.font = hdr_font; c.border = border

for i, r in enumerate(runs_sorted, 4):
    ws4.cell(row=i, column=1, value=r["model"]).border = border
    ws4.cell(row=i, column=2, value=f"{r['score']['n_match']}/17").border = border
    for j, v in enumerate(vols, 3):
        c = ws4.cell(row=i, column=j, value=round(r["cost_usd"] * v, 2))
        c.number_format = '"$"#,##0.00'; c.border = border

for col, w in enumerate([38, 11, 14, 14, 16, 18, 22], 1):
    ws4.column_dimensions[get_column_letter(col)].width = w

wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Sheets: Summary | Per-field | Critical fields | Cost projection")
